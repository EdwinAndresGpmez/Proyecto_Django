from django.shortcuts import get_object_or_404, render, redirect
from django.http import JsonResponse, HttpResponse
from django.urls import reverse
from django.contrib import messages
from django.db import IntegrityError
from django.core.paginator import Paginator
from django.utils.timezone import now

# Módulos para procesar archivos Excel y PDF
import openpyxl
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import cm
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Paragraph, Table, TableStyle
from reportlab.lib.enums import TA_CENTER
from reportlab.lib import colors
import xlrd
from datetime import datetime, time

# Importa tus formularios
from .forms import (
    FormProfesional,
    FormLugares,
    FormHoras,
    PacienteForm,
    FormServicios,
    CargarHorarioArchivoForm,
)

# Importa tus modelos
from .models import (
    Horas,
    Lugares,
    Pacientes,
    Consultorio,
    Auditoria,
    Profesional,
    Servicio,
)

# Importa las referencias de otras apps (si las necesitas)
from usuario import models as modelsUsuario
from entrarSistema import forms as formEntrarSistema
from entrarSistema import models as modelsEntrarSistema
# Django Auth y demás
from functools import wraps



# -------------------------------------------------------------------
# 1. DECORADOR PERSONALIZADO
# -------------------------------------------------------------------
def auditoria_login_required(descripcion=None):
    """
    Decorador que:
      1. Obliga a que el usuario esté autenticado.
      2. Registra en Auditoria el acceso a la vista, guardando una descripción.
    """
    def decorator(view_func):
        @wraps(view_func)
        def _wrapped_view(request, *args, **kwargs):
            # Verificar autenticación
            if not request.user.is_authenticated:
                return redirect('iniciarSesion')

            # Registrar en Auditoria
            texto = descripcion or f"Acceso a la vista {view_func.__name__}"
            Auditoria.objects.create(
                descripcion_aut=f"{texto}. Usuario ID: {request.user.id}. Fecha: {now()}"
            )
            return view_func(request, *args, **kwargs)
        return _wrapped_view
    return decorator


# -------------------------------------------------------------------
# 2. FUNCIÓN AUXILIAR PARA ROLES
# -------------------------------------------------------------------
def TipoRol(request):
    """
    Devuelve la instancia de roles del usuario actual para verificar permisos,
    por ejemplo: tipoRol.es_administrador, tipoRol.es_programador, etc.
    """
    instanceHoras = formEntrarSistema.UsuarioRoles.objects.get(id_usu=request.user)
    return instanceHoras

# -------------------------------------------------------------------
# 3. VISTAS (con el nuevo decorador)
# -------------------------------------------------------------------

@auditoria_login_required(descripcion="Carga masiva de horarios")
def cargar_horario(request):
    if request.method == "POST":
        form = CargarHorarioArchivoForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = request.FILES.get("archivo")
            if archivo:
                if archivo.name.endswith(".xlsx"):
                    try:
                        workbook = openpyxl.load_workbook(archivo)
                        sheet = workbook.active

                        for fila_num, row in enumerate(sheet.iter_rows(values_only=True), start=1):
                            try:
                                if len(row) < 6:
                                    messages.error(request, f"Fila {fila_num} con columnas insuficientes.")
                                    continue

                                id_prof = int(row[0])
                                hora_inicio = row[1]
                                hora_final = row[2]
                                fecha_habilitada = row[3].date()
                                estado_hora = row[4].strip().lower() == "activo"

                                # Obtener profesional
                                try:
                                    profesional = Profesional.objects.get(id_prof=id_prof)
                                except Profesional.DoesNotExist:
                                    messages.error(
                                        request,
                                        f"El profesional con ID {id_prof} no existe."
                                    )
                                    continue

                                # Verificar si ya existe
                                if Horas.objects.filter(
                                    id_prof=profesional,
                                    inicio_hora=hora_inicio,
                                    final_hora=hora_final,
                                    fecha_habilitada=fecha_habilitada
                                ).exists():
                                    messages.warning(
                                        request,
                                        f"Registro duplicado para profesional ID {id_prof}, hora {hora_inicio}-{hora_final}."
                                    )
                                    continue

                                # Crear o actualizar
                                Horas.objects.update_or_create(
                                    id_prof=profesional,
                                    inicio_hora=hora_inicio,
                                    final_hora=hora_final,
                                    fecha_habilitada=fecha_habilitada,
                                    defaults={"horas_estado": estado_hora},
                                )
                                messages.success(
                                    request,
                                    f"Registro actualizado para ID {id_prof}, Hora {hora_inicio}-{hora_final}."
                                )
                            except ValueError as ve:
                                messages.error(request, f"Error de valor en fila {fila_num}: {ve}")
                            except Exception as e:
                                messages.error(request, f"Error al procesar fila {fila_num}: {e}")
                        return render(request, "cargar_horario.html", {"form": form})
                    except Exception as e:
                        messages.error(request, f"Error al procesar el archivo: {str(e)}")
                else:
                    messages.error(request, "El archivo no es XLSX.")
            else:
                messages.error(request, "No se recibió ningún archivo.")
        else:
            messages.error(request, "Formulario inválido.")
    else:
        form = CargarHorarioArchivoForm()
    return render(request, "horarios.html", {"form": form})


@auditoria_login_required(descripcion="Vista de inicio admin")
def inicioAdmin(request):
    # Verificar rol
    tipoRol = TipoRol(request)
    if not tipoRol.es_administrador:
        return redirect("inicio")

    lista_citas = (
        Consultorio.objects.filter(id_cit__citas_estado=True)
        .exclude(id_cit__estado_cita="Rechazada")
        .distinct("id_cit")
        .order_by("-id_cit")
    )
    listaRoles = modelsEntrarSistema.UsuarioRoles.objects.filter(id_usu=request.user.id)
    return render(request, "inicio_admin.html", {"listaCitas": lista_citas, "listaRoles": listaRoles})


@auditoria_login_required(descripcion="Generar PDF de citas")
def pdfCitas(request):
    response = HttpResponse(content_type="application/pdf")
    response["Content-Disposition"] = "attachment; filename=Lista-Citas-reporte.pdf"

    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)

    # Cabecera
    c.setLineWidth(0.3)
    c.setFont("Helvetica", 22)
    c.drawString(30, 745, "Citas")
    c.setFont("Helvetica", 12)
    c.drawString(30, 725, "Reporte de citas realizadas")

    hoy_fecha = str(datetime.today().strftime("%Y-%m-%d"))
    hoy = "Fecha de Hoy: " + hoy_fecha

    c.setFont("Helvetica-Bold", 12)
    c.drawString(420, 750, hoy)
    c.line(415, 745, 570, 745)

    # Encabezados
    styles = getSampleStyleSheet()
    styleBH = styles["Normal"]
    styleBH.alignment = TA_CENTER
    styleBH.fontSize = 10

    id_cita = Paragraph("#", styleBH)
    representante = Paragraph("REPRESENTANTE", styleBH)
    paciente = Paragraph("PACIENTE", styleBH)
    lugar = Paragraph("LUGAR", styleBH)
    fecha = Paragraph("FECHA", styleBH)
    estado = Paragraph("ESTADO", styleBH)

    data = [[id_cita, representante, paciente, lugar, fecha, estado]]

    # Consulta de ejemplo
    lista_consulta = (
        Consultorio.objects.filter(id_cit__estado_cita="Realizada")
        .distinct("id_cit")
        .order_by("-id_cit")
    )

    # Contenido
    high = 600
    for consulta in lista_consulta:
        if consulta.id_cit.id_pac:
            pac = consulta.id_cit.id_pac.nombre_pac
        else:
            pac = "No enlazado"

        id_cit_ = str(consulta.id_cit.id_cit)
        rep = str(consulta.id_cit.id_usu.nombre)
        pac_n = str(pac)
        lug = str(consulta.id_cit.id_lugar.ubicacion_lugar)
        fec = str(consulta.id_cit.dia_cit)
        est = str(consulta.id_cit.estado_cita)

        row_data = [
            Paragraph(id_cit_, styleBH),
            Paragraph(rep, styleBH),
            Paragraph(pac_n, styleBH),
            Paragraph(lug, styleBH),
            Paragraph(fec, styleBH),
            Paragraph(est, styleBH),
        ]
        data.append(row_data)
        high -= 18

    table = Table(
        data,
        colWidths=[1.5*cm, 4.5*cm, 4.5*cm, 3.9*cm, 2.5*cm, 2.5*cm],
    )
    table.setStyle(
        TableStyle(
            [
                ("INNERGRID", (0, 0), (-1, -1), 0.25, colors.black),
                ("BOX", (0, 0), (-1, -1), 0.25, colors.black),
            ]
        )
    )
    table.wrapOn(c, A4[0], A4[1])
    table.drawOn(c, 30, high)
    c.showPage()
    c.save()

    pdf = buffer.getvalue()
    buffer.close()
    response.write(pdf)
    return response


@auditoria_login_required(descripcion="Lista de pacientes")
def lista_pacientes(request):
    if request.method == "POST":
        form = PacienteForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "¡Paciente creado exitosamente!")
            return redirect("lista_pacientes")
        else:
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Error en {field}: {error}")
    else:
        form = PacienteForm()

    pacientes = Pacientes.objects.all()
    paginator = Paginator(pacientes, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    return render(request, "pacientes.html", {"form": form, "page_obj": page_obj})


@auditoria_login_required(descripcion="Eliminar paciente")
def eliminar_paciente(request, id_pac):
    if request.method == "POST":
        paciente = get_object_or_404(Pacientes, id_pac=id_pac)

        citas = modelsUsuario.Citas.objects.filter(id_pac=paciente)
        if citas.exists():
            return JsonResponse({
                "status": "error",
                "message": "Este paciente tiene citas asociadas, no se puede eliminar.",
            })
        paciente.delete()
        return JsonResponse({"status": "success", "message": "Paciente eliminado exitosamente."})
    return JsonResponse({"status": "error", "message": "Error al procesar la solicitud."})


@auditoria_login_required(descripcion="Editar paciente")
def editar_paciente(request, id_pac):
    if request.method == "POST":
        nombre_pac = request.POST.get("nombre_pac")
        nacimiento_pac = request.POST.get("nacimiento_pac")
        genero_pac = request.POST.get("genero_pac")
        estado_pac = request.POST.get("pacientes_estado")

        # Convertir fecha
        nacimiento_pac = datetime.strptime(nacimiento_pac, "%Y-%m-%d").date()
        if nacimiento_pac > datetime.today().date():
            return JsonResponse({
                "status": "error",
                "message": "La fecha de nacimiento no puede ser mayor a hoy.",
            })

        paciente = Pacientes.objects.get(id_pac=id_pac)
        paciente.nombre_pac = nombre_pac
        paciente.nacimiento_pac = nacimiento_pac
        paciente.genero_pac = genero_pac
        paciente.pacientes_estado = True if estado_pac == "True" else False
        paciente.save()

        return JsonResponse({"status": "success"})
    else:
        paciente = Pacientes.objects.get(id_pac=id_pac)
        formatted_nacimiento = paciente.nacimiento_pac.strftime("%Y-%m-%d")
        return render(request, "editar_paciente.html", {
            "paciente": paciente,
            "formatted_nacimiento": formatted_nacimiento,
        })


def is_ajax(request):
    return request.META.get("HTTP_X_REQUESTED_WITH") == "XMLHttpRequest"


@auditoria_login_required(descripcion="Vista consultorio cita")
def consultorio_cita(request):
    # Rol admin
    tipoRol = TipoRol(request)
    if not tipoRol.es_administrador:
        return redirect("inicio")
    # GET
    if request.method == "GET":
        lista_select = modelsUsuario.Citas.objects.filter(id_cit=request.GET["id_cit"])
        consultas = [cita.id_pac for cita in lista_select]
        if None in consultas:
            lista_consulta = ""
        else:
            lista_consulta = Consultorio.objects.filter(
                id_cit__id_pac=consultas[0], consultorio_estado=True
            )
        return render(
            request, "consultorio_citas.html",
            {"citaSelect": lista_select, "listaPrevia": lista_consulta},
        )
    # POST
    instance_Citas = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])
    nacimiento_hidden = request.POST.get("nacimiento_con")
    if nacimiento_hidden and nacimiento_hidden != "nada":
        try:
            nacimiento_convertido = datetime.strptime(nacimiento_hidden, "%d de %B de %Y").date()
        except ValueError:
            return render(
                request,
                "consultorio_citas.html",
                {
                    "error": "El formato de la fecha de nacimiento no es válido. Use YYYY-MM-DD.",
                    "citaSelect": modelsUsuario.Citas.objects.filter(id_cit=request.POST["id_cit"]),
                },
            )
    else:
        nacimiento_convertido = instance_Citas.id_pac.nacimiento_pac

    idCon = Consultorio(
        id_cit=instance_Citas,
        id_prof=instance_Citas.id_prof,
        nota_con=request.POST["nota_con"],
        nacimiento_con=nacimiento_convertido,
    )
    idCon.save()

    idCit = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])
    idCit.estado_cita = "Realizada"
    idCit.save()

    # Registrar auditoría adicional (acción específica)
    Auditoria.objects.create(
        descripcion_aut=f"Se creó una consulta en *Consultorio*, nota:{request.POST['nota_con']}, fecha:{nacimiento_convertido}, profesional:{idCon.id_prof}, id:{idCon.pk}, usuario:{request.user.id}"
    )

    return redirect("consultorio")


@auditoria_login_required(descripcion="Vista consultorio")
def consultorio(request):
    tipoRol = TipoRol(request)
    if not tipoRol.es_administrador:
        return redirect("inicio")

    fecha_hoy = datetime.today().strftime("%Y-%m-%d")
    if request.method == "GET":
        lista_citas = modelsUsuario.Citas.objects.filter(
            citas_estado=True, estado_cita="Aceptada", dia_cit=datetime.today()
        ).order_by("-id_cit")
        lista_consultorio = modelsUsuario.Citas.objects.filter(
            citas_estado=True, estado_cita="Aceptada"
        )
        return render(
            request,
            "consultorio.html",
            {
                "listaCitas": lista_citas,
                "citaSelect": modelsUsuario.Citas.objects.filter(id_cit=9),
                "listaConsultorio": lista_consultorio,
                "fecha": fecha_hoy,
            },
        )
    else:
        lista_citas = modelsUsuario.Citas.objects.filter(citas_estado=True, estado_cita="Aceptada")
        lista_consultorio = modelsUsuario.Citas.objects.filter(citas_estado=True, estado_cita="Aceptada")

        instance_Citas = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])
        idCon = Consultorio(id_cit=instance_Citas)
        idCon.save()

        idCit = modelsUsuario.Citas(id_cit=request.POST["id_cit"])
        idCit.estado_cita = "Realizada"
        idCit.save()

        Auditoria.objects.create(
            descripcion_aut=f"Se creó una 'consulta' en *Consultorio*, nota:{request.POST['nota_con']}, id:{idCon.pk}, usuario:{request.user.id}"
        )

        return render(
            request,
            "consultorio.html",
            {
                "listaCitas": lista_citas,
                "citaSelect": modelsUsuario.Citas.objects.filter(id_cit=9),
                "listaConsultorio": lista_consultorio,
                "fecha": fecha_hoy,
            },
        )


@auditoria_login_required(descripcion="Vista de citas admin")
def citas_admin(request):
    tipoRol = TipoRol(request)
    if not tipoRol.es_administrador:
        return redirect("inicio")

    if request.method == "GET":
        lista_citas_aceptadas = modelsUsuario.Citas.objects.filter(estado_cita="Aceptada", citas_estado=True)
        lista_citas_rechazadas = modelsUsuario.Citas.objects.filter(estado_cita="Rechazada", citas_estado=False)
        lista_citas_pendientes = modelsUsuario.Citas.objects.filter(estado_cita="Sin confirmar", citas_estado=True)
        lista_citas_realizadas = modelsUsuario.Citas.objects.filter(estado_cita="Realizada", citas_estado=True)
        lista_citas_canceladas = modelsUsuario.Citas.objects.filter(estado_cita="Cancelada", citas_estado=False)

        return render(
            request,
            "citas_admin.html",
            {
                "listaCitasAceptadas": lista_citas_aceptadas,
                "listaCitasRechazadas": lista_citas_rechazadas,
                "listaCitasPendientes": lista_citas_pendientes,
                "listaCitasRealizadas": lista_citas_realizadas,
                "listaCitasCancelada": lista_citas_canceladas,
            },
        )
    else:
        lista_citas_aceptadas = modelsUsuario.Citas.objects.filter(estado_cita="Aceptada", citas_estado=True)
        lista_citas_rechazadas = modelsUsuario.Citas.objects.filter(estado_cita="Rechazada", citas_estado=False)
        lista_citas_pendientes = modelsUsuario.Citas.objects.filter(estado_cita="Sin confirmar", citas_estado=True)
        lista_citas_canceladas = modelsUsuario.Citas.objects.filter(estado_cita="Cancelada", citas_estado=False)

        if "btnRevisarAceptar" in request.POST:
            idCita = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])
            idCita.estado_cita = "Aceptada"
            idCita.save()
            Auditoria.objects.create(
                descripcion_aut=f"Se aceptó cita ID {request.POST['id_cit']}, usuario:{request.user.id}"
            )

        elif "btnRevisarRechazar" in request.POST:
            idCita = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])
            idCita.estado_cita = "Rechazada"
            idCita.citas_estado = False
            idCita.save()
            Auditoria.objects.create(
                descripcion_aut=f"Se rechazó cita ID {request.POST['id_cit']}, usuario:{request.user.id}"
            )

        elif "btnAceptadaRechazar" in request.POST:
            idCita = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])
            idCita.estado_cita = "Rechazada"
            idCita.citas_estado = False
            idCita.save()
            Auditoria.objects.create(
                descripcion_aut=f"Se rechazó (antes aceptada) cita ID {request.POST['id_cit']}, usuario:{request.user.id}"
            )

        elif "btnRechazarAceptar" in request.POST:
            idCita = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])
            idCita.estado_cita = "Aceptada"
            idCita.citas_estado = True
            idCita.save()
            Auditoria.objects.create(
                descripcion_aut=f"Se aceptó (antes rechazada) cita ID {request.POST['id_cit']}, usuario:{request.user.id}"
            )

        elif "btnRechazarEliminar" in request.POST:
            idCita = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])
            idCita.citas_estado = False
            idCita.save()
            Auditoria.objects.create(
                descripcion_aut=f"Se eliminó cita ID {request.POST['id_cit']}, usuario:{request.user.id}"
            )

        elif "btnRevisarCancelar" in request.POST:
            idCita = modelsUsuario.Citas.objects.get(id_cit=request.POST["id_cit"])
            idCita.citas_estado = False
            idCita.estado_cita = "Cancelada"
            idCita.save()
            Auditoria.objects.create(
                descripcion_aut=f"Se canceló cita ID {request.POST['id_cit']}, usuario:{request.user.id}"
            )

        return render(
            request,
            "citas_admin.html",
            {
                "listaCitasAceptadas": lista_citas_aceptadas,
                "listaCitasRechazadas": lista_citas_rechazadas,
                "listaCitasPendientes": lista_citas_pendientes,
                "listaCitasCancelada": lista_citas_canceladas,
            },
        )


@auditoria_login_required(descripcion="Configuración admin")
def cfg_admin(request):
    tipoRol = TipoRol(request)
    if not tipoRol.es_usuario:
        return redirect("inicio")

    form = formEntrarSistema.FormRegistrar()
    if request.method == "GET":
        return render(request, "configuracion_admin.html", {"form": form})
    else:
        idUser = modelsEntrarSistema.CrearCuenta.objects.get(id=request.user.id)

        if "btnUsuario" in request.POST:
            if modelsEntrarSistema.CrearCuenta.objects.filter(username=request.POST["username"]).exists():
                messages.error(request, "Ya existe ese usuario.")
                return render(request, "configuracion_admin.html", {"form": form})
            else:
                idUser.username = request.POST["username"]
                idUser.save()
                Auditoria.objects.create(
                    descripcion_aut=f"Usuario {request.user.id} modificó su 'username' a {request.POST['username']}"
                )
                return render(request, "configuracion_admin.html", {"form": form})

        elif "btnCedula" in request.POST:
            if modelsEntrarSistema.CrearCuenta.objects.filter(cedula=request.POST["cedula"]).exists():
                messages.error(request, "Ya se ha registrado esta cédula.")
                return render(request, "configuracion_admin.html", {"form": form})
            else:
                idUser.cedula = request.POST["cedula"]
                idUser.save()
                Auditoria.objects.create(
                    descripcion_aut=f"Usuario {request.user.id} modificó su 'cedula' a {request.POST['cedula']}"
                )
                return render(request, "configuracion_admin.html", {"form": form})

        elif "btnTelefono" in request.POST:
            if not request.POST["numero"]:
                messages.error(request, "Tiene que escribir un número telefónico.")
                return render(request, "configuracion_admin.html", {"form": form})
            else:
                idUser.numero = request.POST["numero"]
                idUser.save()
                Auditoria.objects.create(
                    descripcion_aut=f"Usuario {request.user.id} modificó su 'número' a {request.POST['numero']}"
                )
                return render(request, "configuracion_admin.html", {"form": form})

        elif "btnCorreo" in request.POST:
            if modelsEntrarSistema.CrearCuenta.objects.filter(correo=request.POST["correo"]).exists():
                messages.error(request, "Ya existe este correo.")
                return render(request, "configuracion_admin.html", {"form": form})
            else:
                idUser.correo = request.POST["correo"]
                idUser.save()
                Auditoria.objects.create(
                    descripcion_aut=f"Usuario {request.user.id} modificó su 'correo' a {request.POST['correo']}"
                )
                return render(request, "configuracion_admin.html", {"form": form})

        elif "btnContrasenna" in request.POST:
            if request.POST["password1"] != request.POST["password2"]:
                messages.error(request, "Las contraseñas deben ser iguales.")
                return render(request, "configuracion_admin.html", {"form": form})
            else:
                idUser.set_password(request.POST["password1"])
                idUser.save()
                Auditoria.objects.create(
                    descripcion_aut=f"Usuario {request.user.id} modificó su 'contraseña'."
                )
                return render(request, "configuracion_admin.html", {"form": form})

        elif "btnAll" in request.POST:
            if (
                modelsEntrarSistema.CrearCuenta.objects.filter(correo=request.POST["correo"]).exists() or
                modelsEntrarSistema.CrearCuenta.objects.filter(username=request.POST["username"]).exists() or
                modelsEntrarSistema.CrearCuenta.objects.filter(cedula=request.POST["cedula"]).exists()
            ):
                messages.error(request, "Algunos datos ya existen en el sistema, elige otros por favor.")
                return render(request, "configuracion_admin.html", {"form": form})
            else:
                if request.POST["password1"] == request.POST["password2"]:
                    idUser.username = request.POST["username"]
                    idUser.cedula = request.POST["cedula"]
                    idUser.numero = request.POST["numero"]
                    idUser.correo = request.POST["correo"]
                    idUser.set_password(request.POST["password1"])
                    idUser.save()

                    Auditoria.objects.create(
                        descripcion_aut=f"Usuario {request.user.id} modificó todos sus datos."
                    )
                    return render(request, "configuracion_admin.html", {"form": form})
                else:
                    messages.error(request, "Las contraseñas deben ser iguales.")
                    return render(request, "configuracion_admin.html", {"form": form})

    return render(request, "configuracion_admin.html", {"form": form})


# -------------------------------------------------------------------
# PROFESIONALES
# -------------------------------------------------------------------
@auditoria_login_required(descripcion="Lista de profesionales")
def lista_profesionales(request):
    if request.method == "POST":
        if "submitProfesional" in request.POST:
            formProfesional = FormProfesional(request.POST)
            if formProfesional.is_valid():
                try:
                    nuevo_profesional = formProfesional.save(commit=False)
                    lugares_seleccionados = formProfesional.cleaned_data.get("lugares")

                    # Verificar que no haya choque de lugares
                    lugares_ocupados = Profesional.objects.filter(
                        lugares__in=lugares_seleccionados
                    ).distinct()
                    if lugares_ocupados.exists():
                        lugar_nombres = ", ".join(
                            lugar for lugar in lugares_ocupados.values_list(
                                "lugares__nombre_lugar", flat=True
                            )
                        )
                        formProfesional.add_error(
                            "lugares",
                            f"Los lugares '{lugar_nombres}' ya están asociados a otros profesionales."
                        )
                        return render(
                            request,
                            "profesionales.html",
                            {
                                "formProfesional": formProfesional,
                                "profesionales": Profesional.objects.all(),
                                "lugares": Lugares.objects.all(),
                            },
                        )

                    nuevo_profesional.save()
                    if lugares_seleccionados:
                        nuevo_profesional.lugares.set(lugares_seleccionados)
                    return redirect("lista_profesionales")
                
                except IntegrityError:
                    formProfesional.add_error(
                        "num_doc_prof", "El número de documento ya existe."
                    )
            else:
                return render(
                    request,
                    "profesionales.html",
                    {
                        "formProfesional": formProfesional,
                        "profesionales": Profesional.objects.all(),
                        "lugares": Lugares.objects.all(),
                    },
                )

    formProfesional = FormProfesional()
    profesionales = Profesional.objects.all()
    lugares = Lugares.objects.all()
    lugares_ids = lugares.values_list("id_lugar", flat=True)
    return render(
        request,
        "profesionales.html",
        {
            "formProfesional": formProfesional,
            "profesionales": profesionales,
            "lugares": lugares,
            "lugares_ids": lugares_ids,
        },
    )


@auditoria_login_required(descripcion="Editar profesional")
def editar_profesional(request, id_prof):
    if request.method == "POST":
        profesional = get_object_or_404(Profesional, id_prof=id_prof)

        profesional.nombre_prof = request.POST.get("nombre_prof", profesional.nombre_prof)
        profesional.especialidad_prof = request.POST.get("especialidad_prof", profesional.especialidad_prof)
        profesional.email_prof = request.POST.get("email_prof", profesional.email_prof)
        profesional.telefono_prof = request.POST.get("telefono_prof", profesional.telefono_prof)
        profesional.estado_prof = request.POST.get("estado_prof") == "True"

        lugares_ids = request.POST.getlist("lugares[]")
        if lugares_ids:
            lugares_set = Lugares.objects.filter(id_lugar__in=lugares_ids)
            # Verificar si algún otro profesional ya está asociado al mismo lugar
            for lugar in lugares_set:
                otros_profesionales = lugar.profesionales.exclude(id_prof=id_prof)
                if otros_profesionales.exists():
                    return JsonResponse({
                        "success": False,
                        "message": f"El lugar '{lugar.nombre_lugar}' ya está asociado a otro profesional.",
                    })
            profesional.lugares.set(lugares_set)
        else:
            profesional.lugares.clear()

        profesional.save()
        return JsonResponse({"success": True, "message": "Profesional editado correctamente."})
    else:
        return JsonResponse({"status": "error", "message": "Método no permitido."})


@auditoria_login_required(descripcion="Eliminar profesional")
def eliminar_profesional(request, id_prof):
    if request.method == "POST":
        profesional = get_object_or_404(Profesional, id_prof=id_prof)

        citas = modelsUsuario.Citas.objects.filter(id_prof=profesional)
        if citas.exists():
            return JsonResponse({
                "status": "error",
                "message": "Este profesional tiene citas asociadas, no se puede eliminar."
            })

        horas = Horas.objects.filter(id_prof=profesional)
        if horas.exists():
            return JsonResponse({
                "status": "error",
                "message": "Este profesional tiene horarios asociados, no se puede eliminar."
            })

        try:
            profesional.delete()
            return JsonResponse({"status": "success", "message": "Profesional eliminado exitosamente."})
        except IntegrityError as e:
            return JsonResponse({"status": "error", "message": f"Error de integridad: {str(e)}"})
        except Exception as e:
            return JsonResponse({"status": "error", "message": f"Error al eliminar al profesional: {str(e)}"})

    return JsonResponse({"status": "error", "message": "Error al procesar la solicitud."})


# -------------------------------------------------------------------
# LUGARES
# -------------------------------------------------------------------
@auditoria_login_required(descripcion="Lista de lugares")
def lista_lugar(request):
    if request.method == "POST":
        if "submitLugares" in request.POST:
            formLugares = FormLugares(request.POST)
            if formLugares.is_valid():
                formLugares.save()
                return redirect("lista_lugar")
    else:
        formLugares = FormLugares()
        lugares = Lugares.objects.all()

    context = {
        "formLugares": formLugares,
        "lugares": lugares,
    }
    return render(request, "lugares.html", context)


@auditoria_login_required(descripcion="Editar lugar")
def editar_lugar(request, id_lugar):
    if request.method == "POST":
        lugar = Lugares.objects.get(id_lugar=id_lugar)
        lugar.nombre_lugar = request.POST["nombre_lugar"]
        lugar.ubicacion_lugar = request.POST["ubicacion_lugar"]
        lugar.lugares_estado = request.POST["lugares_estado"]
        lugar.save()
        return JsonResponse({"status": "success"})
    return JsonResponse({"status": "error"})


@auditoria_login_required(descripcion="Eliminar lugar")
def eliminar_lugar(request, id_lugar):
    if request.method == "POST":
        try:
            lugar = Lugares.objects.get(id_lugar=id_lugar)
            if modelsUsuario.Citas.objects.filter(id_lugar=lugar).exists():
                return JsonResponse({
                    "status": "error",
                    "message": "No se puede eliminar el lugar, está asociado a citas.",
                })
            if Profesional.objects.filter(lugares=lugar).exists():
                return JsonResponse({
                    "status": "error",
                    "message": "No se puede eliminar el lugar, está asociado a profesionales.",
                })
            lugar.delete()
            return JsonResponse({"status": "success", "message": "Lugar eliminado exitosamente."})
        except Lugares.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Lugar no encontrado."})


# -------------------------------------------------------------------
# HORARIOS
# -------------------------------------------------------------------
@auditoria_login_required(descripcion="Lista de horarios")
def lista_horario(request):
    formHoras = FormHoras()
    query = Horas.objects.all()

    if request.method == "POST":
        if "archivoExcel" in request.FILES:
            # Procesar archivo, etc.
            archivo = request.FILES["archivoExcel"]
            try:
                messages.success(request, "Archivo cargado correctamente.")
            except Exception as e:
                messages.error(request, f"Error al procesar el archivo: {str(e)}")
            return redirect("lista_horario")

        elif "submitHoras" in request.POST:
            formHoras = FormHoras(request.POST)
            if formHoras.is_valid():
                formHoras.save()
                messages.success(request, "Horario creado exitosamente.")
                return redirect("lista_horario")
            else:
                messages.error(request, "Corrige los errores en el formulario.")

        elif "submitHorasFiltro" in request.POST:
            if "num_doc_prof" in request.POST and request.POST["num_doc_prof"]:
                query = query.filter(
                    id_prof__num_doc_prof__icontains=request.POST["num_doc_prof"]
                )
            if "fecha_habilitada" in request.POST and request.POST["fecha_habilitada"]:
                query = query.filter(fecha_habilitada=request.POST["fecha_habilitada"])
            if "estado_hora" in request.POST and request.POST["estado_hora"]:
                estado_hora = request.POST["estado_hora"]
                if estado_hora == "Activo":
                    query = query.filter(horas_estado=True)
                elif estado_hora == "Inactivo":
                    query = query.filter(horas_estado=False)

    paginator = Paginator(query, 10)
    page_number = request.GET.get("page")
    page_obj = paginator.get_page(page_number)

    context = {
        "formHoras": formHoras,
        "page_obj": page_obj,
    }
    return render(request, "horarios.html", context)


@auditoria_login_required(descripcion="Editar horario")
def editar_horario(request, id_hora):
    if request.method == "POST":
        try:
            horario = Horas.objects.get(id_hora=id_hora)
        except Horas.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Horario no encontrado"})

        horario.inicio_hora = request.POST.get("inicio_hora", horario.inicio_hora)
        horario.final_hora = request.POST.get("final_hora", horario.final_hora)

        estado = request.POST.get("horas_estado")
        if estado == "A":
            horario.horas_estado = True
        elif estado == "I":
            horario.horas_estado = False
        else:
            return JsonResponse({"status": "error", "message": "Estado inválido"}, status=400)

        try:
            fecha_habilitada = request.POST.get("fecha_habilitada")
            if fecha_habilitada:
                horario.fecha_habilitada = datetime.strptime(fecha_habilitada, "%Y-%m-%d").date()
        except ValueError:
            return JsonResponse({"status": "error", "message": "Fecha habilitada inválida"})

        try:
            profesional = Profesional.objects.get(num_doc_prof=request.POST.get("num_doc_prof"))
            horario.id_prof = profesional
        except Profesional.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Profesional no encontrado"})

        horario.save()
        return JsonResponse({"status": "success", "message": "Horario actualizado exitosamente."})
    return JsonResponse({"status": "error", "message": "Método no permitido"})


@auditoria_login_required(descripcion="Eliminar horario")
def eliminar_horario(request, id_hora):
    if request.method == "POST":
        try:
            horario = Horas.objects.get(id_hora=id_hora)
            citas = modelsUsuario.Citas.objects.filter(id_hora=horario)
            if citas.exists():
                return JsonResponse({
                    "status": "error",
                    "message": "Este horario está asociado a citas, no se puede eliminar.",
                })
            horario.delete()
            return JsonResponse({"status": "success", "message": "Horario eliminado exitosamente."})
        except Horas.DoesNotExist:
            return JsonResponse({"status": "error", "message": "Horario no encontrado"})
    return JsonResponse({"status": "error", "message": "Error al procesar la solicitud."})


# -------------------------------------------------------------------
# SERVICIOS
# -------------------------------------------------------------------
@auditoria_login_required(descripcion="Lista de servicios")
def lista_servicios(request):
    if request.method == "POST":
        formServicios = FormServicios(request.POST)
        if formServicios.is_valid():
            nombre_servicio = formServicios.cleaned_data.get("nombre_servicio")
            if Servicio.objects.filter(nombre_servicio=nombre_servicio).exists():
                messages.error(
                    request,
                    f"El servicio '{nombre_servicio}' ya existe. Elige otro nombre.",
                )
            else:
                servicio = formServicios.save(commit=False)
                profesionales = formServicios.cleaned_data.get("profesionales")
                servicio.save()
                servicio.profesionales.set(profesionales)
                messages.success(request, "Servicio creado exitosamente.")
                return redirect("lista_servicios")
        else:
            messages.error(request, "Corrige los errores del formulario.")
    else:
        formServicios = FormServicios()

    servicios = Servicio.objects.prefetch_related("profesionales").all()
    profesionales = Profesional.objects.all()
    return render(request, "servicios.html", {
        "formServicios": formServicios,
        "servicios": servicios,
        "profesionales": profesionales,
    })


@auditoria_login_required(descripcion="Eliminar servicio")
def eliminar_servicios(request, id_servicio):
    if request.method == "POST":
        servicio = get_object_or_404(Servicio, id_servicio=id_servicio)
        citas = modelsUsuario.Citas.objects.filter(id_servicio=servicio)
        if citas.exists():
            return JsonResponse({
                "status": "error",
                "message": "Este servicio tiene citas asociadas, no se puede eliminar.",
            })
        servicio.delete()
        return JsonResponse({"status": "success", "message": "Servicio eliminado exitosamente."})
    return JsonResponse({"status": "error", "message": "Error al procesar la solicitud."})


@auditoria_login_required(descripcion="Editar servicio")
def editar_servicios(request, id_servicio):
    if request.method == "POST":
        servicio = get_object_or_404(Servicio, id_servicio=id_servicio)
        servicio.nombre_servicio = request.POST.get("nombre_servicio", servicio.nombre_servicio)
        servicio.descripcion_servicio = request.POST.get("descripcion_servicio", servicio.descripcion_servicio)
        servicio.servicio_estado = request.POST.get("servicio_estado") == "True"

        profesionales_ids = request.POST.getlist("profesionales[]")
        if profesionales_ids:
            profesionales = Profesional.objects.filter(id_prof__in=profesionales_ids)
            servicio.profesionales.set(profesionales)
        else:
            servicio.profesionales.clear()

        servicio.save()
        servicio_actualizado = {
            "id_servicio": servicio.id_servicio,
            "nombre_servicio": servicio.nombre_servicio,
            "descripcion_servicio": servicio.descripcion_servicio,
            "servicio_estado": servicio.servicio_estado,
            "profesionales": [
                {"id_prof": prof.id_prof, "nombre_prof": prof.nombre_prof}
                for prof in servicio.profesionales.all()
            ]
        }
        return JsonResponse({
            "status": "success",
            "message": "Servicio editado correctamente.",
            "servicio": servicio_actualizado
        })
    return JsonResponse({"status": "error", "message": "Método no permitido."})


# -------------------------------------------------------------------
# REPORTES
# -------------------------------------------------------------------
@auditoria_login_required(descripcion="Generar reporte XLSX")
def generate_report(request, report_type):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={report_type}_report.xlsx'
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte"

    if report_type == "citas_por_estado":
        ws.append(["ID Cita", "Paciente", "Estado", "Fecha", "Profesional", "Servicio"])
        for cita in modelsUsuario.Citas.objects.all():
            ws.append([
                cita.id_cit,
                cita.id_pac.nombre_pac if cita.id_pac else "Sin paciente",
                cita.estado_cita,
                cita.dia_cit,
                cita.id_prof.nombre_prof if cita.id_prof else "Sin profesional",
                cita.id_servicio.nombre_servicio if cita.id_servicio else "Sin servicio",
            ])
    elif report_type == "profesionales_mas_demandados":
        ws.append(["Profesional", "Número de Citas"])
        for profesional in Profesional.objects.all():
            num_citas = modelsUsuario.Citas.objects.filter(id_prof=profesional).count()
            ws.append([profesional.nombre_prof, num_citas])
    elif report_type == "pacientes_activos":
        ws.append(["ID Paciente", "Nombre", "Tipo Documento", "Número Documento", "Estado"])
        for paciente in Pacientes.objects.filter(pacientes_estado=True):
            ws.append([
                paciente.id_pac,
                paciente.nombre_pac,
                paciente.get_tipo_doc_display(),
                paciente.num_doc,
                "Activo",
            ])
    else:
        ws.append(["Error", "Tipo de reporte no reconocido"])

    wb.save(response)
    return response


@auditoria_login_required(descripcion="Lista de reportes")
def reportes_list(request):
    return render(request, 'report_list.html')
